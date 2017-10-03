import img
from openpyxl import load_workbook

courses = ['1 курс', '2 курс', '3 курс', '4 курс']
kurs4 = {'4КСК7': 4, '4ПКС8': 6, '4АК4': 8, '4Т7': 10, '4Д2': 12, '4ПП2': 14}
kurs3 = {'3КСК8': 4, '3КСК9': 6, '3ПКС9': 8, '3ПКС10': 10, '3АК6': 12, '3Ф6': 14, '3Т9': 16, '3Т10': 18, '3ОДЛ6': 20, '3ИД1': 2, '3Д3': 24, '3ПП3': 26}
days = {0: 'Понедельник', 1: 'Вторник', 2: 'Среда', 3: 'Четверг', 4: 'Пятница', 5: 'Суббота'}
kurs2 = {'2КСК10': 3, '2КСК11': 5, '2ПКС11': 7, '2ПКС12': 9, '2АК7': 11, '2Ф7': 13, '2Т11': 15, '2Т12': 17, '2ОДЛ7': 19, '2ИД2': 21, '2Д4': 23}
msg = {'text': ''}
msg['text'] = 'Гем'
lessons = msg['text'] + '\n'
aud = '\n'
wb = load_workbook('xlsx/%s.xlsx' % (days[1]))
print(wb.get_sheet_names())
#ws = wb['Расписание (4 курс)']
islesson = False

for r in range(3, 10):
    for c in range(4, 26, 2):
        for ws in wb:
            print(ws.cell(row=r, column=c).value)
            if ws.cell(row=r, column=c).value is not None and msg['text'] in ws.cell(row=r, column=c).value:
                if islesson:
                    aud = aud[:-1] + ' ' + ws.cell(row=2, column=c).value
                else:
                    lessons += ws.cell(row=r, column=c).value.replace('\n', ' / ')
                    islesson = True
                    if ws.cell(row=r, column=c + 1).value is not None:
                        aud += ws.cell(row=r, column=c + 1).value.replace('\n', ' / ') + ' ' + ws.cell(row=2, column=c).value
    lessons += '\n'
    aud += '\n'
    islesson = False


'''
for x in range(3, 10):
    if ws.cell(row=x, column=kurs3[msg['text']]).value is None:
        lessons += '\n'
    else:
        lessons += ws.cell(row=x, column=kurs3[msg['text']]).value.replace('\n', ' / ') + '\n'
    if ws.cell(row=x, column=kurs3[msg['text']] + 1).value is None:
        aud += '\n'
    else:
        aud += ws.cell(row=x, column=kurs3[msg['text']] + 1).value.replace('\n', ' / ') + '\n'
'''
print(lessons)
img.render([lessons, aud], (85, 85, 85), 'test')
