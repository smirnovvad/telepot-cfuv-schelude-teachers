import sys
import asyncio
import telepot.aio
from telepot import glance
from telepot.aio.loop import MessageLoop
from telepot.namedtuple import InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup
from telepot.aio.delegate import (
    per_chat_id, per_callback_query_origin, create_open, pave_event_space)
import datetime
import config
import img
from openpyxl import load_workbook


def render(name, day):
    lessons = name + '    ' + day + '\n'
    aud = '\n'
    try:
        wb = load_workbook('../telepot_cfuv/xlsx/%s.xlsx' % (day))
        # print(wb.get_sheet_names())
        # ws = wb['Расписание (4 курс)']
        islesson = False

        for r in range(3, 10):
            for c in range(4, 26, 2):
                for ws in wb:
                    if ws.cell(row=r, column=c).value is not None and name in ws.cell(row=r, column=c).value:
                        if islesson:
                            aud = aud[:-1] + ' ' + ws.cell(row=2, column=c).value
                        else:
                            lessons += ws.cell(row=r, column=c).value.replace('\n', ' / ')
                            islesson = True
                            if ws.cell(row=r, column=c + 1).value is not None:
                                aud += ws.cell(row=r, column=c + 1).value.replace('\n', ' / ') + ' ' + ws.cell(row=2, column=c).value
            lessons += '\n'
            aud += '\n'
            islesson = False

        img.render([lessons, aud], (85, 85, 85), name + day)

    except Exception as e:
        print(e)


class SheluderStarter(telepot.aio.helper.ChatHandler):
    def __init__(self, *args, **kwargs):
        super(SheluderStarter, self).__init__(*args, **kwargs)

    async def _show_teachers(self, teachers, msg):
        teachers = [elem for elem in teachers if msg in elem]
        await self.sender.sendMessage('Выберите преподавателя', reply_markup=InlineKeyboardMarkup(inline_keyboard=list(map(lambda c: [InlineKeyboardButton(text=str(c), callback_data=str(c))], list(teachers)))))

    async def on_chat_message(self, msg):
        content_type, chat_type, chat_id = glance(msg)
        print('Chat:', content_type, chat_type, msg['text'], datetime.datetime.fromtimestamp(msg['date']).strftime('%Y-%m-%d %H:%M:%S')
              )
        if content_type == 'text':
            if any(msg['text'] in s for s in config.teachers):
                await self._show_teachers(config.teachers, msg['text'])
            else:
                await self.sender.sendMessage(
                    'Напишите Фамилию')
        self.close()  # let Quizzer take over


class Sheluder(telepot.aio.helper.CallbackQueryOriginHandler):
    def __init__(self, *args, **kwargs):
        super(Sheluder, self).__init__(*args, **kwargs)
        self._name = ''
        self._day = ''

    async def _show_days(self, days):
        await self.editor.editMessageText('Выбери день недели',
                                          reply_markup=InlineKeyboardMarkup(
                                              inline_keyboard=list(map(lambda c: [InlineKeyboardButton(text=str(c), callback_data=str(c))], list(days.values())))

                                          )
                                          )

    async def on_callback_query(self, msg):
        query_id, from_id, query_data = glance(msg, flavor='callback_query')
        self.sender = telepot.aio.helper.Sender(self.bot, from_id)
        if query_data in config.teachers:
            self._name = query_data
            await self._show_days(config.days)
        elif query_data in config.days.values():
            self._day = query_data
            render(self._name, self._day)
            print(self._name, self._day)
            await self.sender.sendPhoto(open('img/%s.png' % (self._name + self._day), 'rb'))

    async def on__idle(self, event):
        await asyncio.sleep(5)
        await self.editor.deleteMessage()

        self.close()


def chunks(lst, chunk_count):
    chunk_size = len(lst) // chunk_count
    return [lst[i:i + chunk_size] for i in range(0, len(lst), chunk_size)]


'''
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text='Press me', callback_data='4')],
    ])

    bot.sendMessage(chat_id, 'Use inline keyboard', reply_markup=keyboard)
'''

TOKEN = sys.argv[1]  # get token from command-line

# bot = telepot.aio.Bot(TOKEN)

bot = telepot.aio.DelegatorBot(TOKEN, [
    pave_event_space()(
        per_chat_id(), create_open, SheluderStarter, timeout=3),
    pave_event_space()(
        per_callback_query_origin(), create_open, Sheluder, timeout=60),
])


loop = asyncio.get_event_loop()

loop.create_task(MessageLoop(bot).run_forever())
print('Listening ...')

loop.run_forever()
